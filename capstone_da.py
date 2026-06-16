import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from scipy import stats
import requests
from bs4 import BeautifulSoup
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import warnings
warnings.filterwarnings('ignore')

print("=" * 60)
print("GLOBAL SALES INTELLIGENCE DASHBOARD")
print("=" * 60)

# ──────────────────────────────────────────────
# STEP 1: CREATE SALES DATA
# ──────────────────────────────────────────────
print("\n[1/6] Loading sales data...")

data = {
    'order_id': range(1, 51),
    'product': np.random.choice(['Laptop', 'Phone', 'Tablet', 'Monitor', 'Keyboard'], 50),
    'country': np.random.choice(['Germany', 'India', 'USA', 'UK', 'France'], 50),
    'amount_eur': np.random.randint(200, 2000, 50),
    'month': np.random.choice(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'], 50),
    'units': np.random.randint(1, 10, 50)
}

df = pd.DataFrame(data)
df['revenue'] = df['amount_eur'] * df['units']
print(f"Loaded {len(df)} sales records across {df['country'].nunique()} countries")

# ──────────────────────────────────────────────
# STEP 2: SCRAPE LIVE EXCHANGE RATE
# ──────────────────────────────────────────────
print("\n[2/6] Fetching live EUR/USD exchange rate...")

try:
    url = "https://www.x-rates.com/table/?from=EUR&amount=1"
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers, timeout=5)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    rate = None
    for row in soup.find_all('tr'):
        cells = row.find_all('td')
        if cells and 'US Dollar' in row.text:
            rate = float(cells[1].text.strip())
            break
    
    if not rate:
        rate = 1.08
        print(f"Could not scrape rate, using default: 1 EUR = {rate} USD")
    else:
        print(f"Live rate: 1 EUR = {rate} USD")

except:
    rate = 1.08
    print(f"Using default rate: 1 EUR = {rate} USD")

df['amount_usd'] = (df['amount_eur'] * rate).round(2)
df['revenue_usd'] = (df['revenue'] * rate).round(2)

# ──────────────────────────────────────────────
# STEP 3: STATISTICAL ANALYSIS
# ──────────────────────────────────────────────
print("\n[3/6] Running statistical analysis...")

# Product analysis
product_stats = df.groupby('product').agg(
    total_revenue=('revenue_usd', 'sum'),
    avg_order=('amount_usd', 'mean'),
    total_units=('units', 'sum'),
    num_orders=('order_id', 'count')
).round(2).reset_index().sort_values('total_revenue', ascending=False)

# Country analysis
country_stats = df.groupby('country').agg(
    total_revenue=('revenue_usd', 'sum'),
    avg_order=('amount_usd', 'mean'),
    num_orders=('order_id', 'count')
).round(2).reset_index().sort_values('total_revenue', ascending=False)

# Monthly trend
monthly_stats = df.groupby('month').agg(
    total_revenue=('revenue_usd', 'sum'),
    num_orders=('order_id', 'count')
).round(2).reindex(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'])

# Key stats
total_revenue = df['revenue_usd'].sum()
avg_order = df['amount_usd'].mean()
best_product = product_stats.iloc[0]['product']
best_country = country_stats.iloc[0]['country']
std_dev = df['revenue_usd'].std()

print(f"Total Revenue: ${total_revenue:,.0f}")
print(f"Average Order: ${avg_order:.0f}")
print(f"Best Product: {best_product}")
print(f"Best Country: {best_country}")

# ──────────────────────────────────────────────
# STEP 4: GENERATE CHARTS
# ──────────────────────────────────────────────
print("\n[4/6] Generating charts...")

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle('Global Sales Intelligence Dashboard', fontsize=16, fontweight='bold')

colors = ['#2B86C5', '#1D9E75', '#FF3CAC', '#FFE74C', '#FF6B35']

# Chart 1: Revenue by product
axes[0,0].bar(product_stats['product'], product_stats['total_revenue'],
              color=colors[:len(product_stats)])
axes[0,0].set_title('Total Revenue by Product (USD)')
axes[0,0].set_ylabel('Revenue ($)')
axes[0,0].yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))

# Chart 2: Revenue by country
axes[0,1].barh(country_stats['country'], country_stats['total_revenue'],
               color=colors[:len(country_stats)])
axes[0,1].set_title('Revenue by Country (USD)')
axes[0,1].xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))

# Chart 3: Monthly trend
axes[1,0].plot(monthly_stats.index, monthly_stats['total_revenue'],
               marker='o', color='#2B86C5', linewidth=2, markersize=8)
axes[1,0].fill_between(monthly_stats.index, monthly_stats['total_revenue'],
                        alpha=0.2, color='#2B86C5')
axes[1,0].set_title('Monthly Revenue Trend (USD)')
axes[1,0].set_ylabel('Revenue ($)')
axes[1,0].yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))

# Chart 4: Revenue distribution
axes[1,1].hist(df['revenue_usd'], bins=15, color='#1D9E75', edgecolor='white')
axes[1,1].axvline(df['revenue_usd'].mean(), color='red', linestyle='--',
                  label=f'Mean: ${df["revenue_usd"].mean():,.0f}')
axes[1,1].set_title('Revenue Distribution')
axes[1,1].set_xlabel('Revenue ($)')
axes[1,1].legend()

plt.tight_layout()
plt.savefig('dashboard_charts.png', dpi=150, bbox_inches='tight')
plt.show()
print("Charts saved!")

# ──────────────────────────────────────────────
# STEP 5: GENERATE EXCEL REPORT
# ──────────────────────────────────────────────
print("\n[5/6] Generating Excel report...")

wb = openpyxl.Workbook()

# ── Sheet 1: Raw Data
ws1 = wb.active
ws1.title = "Raw Data"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2B86C5", end_color="2B86C5", fill_type="solid")
center = Alignment(horizontal='center')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = list(df.columns)
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h.upper())
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

for row_idx, row in df.iterrows():
    for col_idx, value in enumerate(row, 1):
        ws1.cell(row=row_idx+2, column=col_idx, value=value)

# ── Sheet 2: Product Summary
ws2 = wb.create_sheet("Product Summary")
ws2['A1'] = "Product Performance Summary"
ws2['A1'].font = Font(bold=True, size=14, color="2B86C5")

headers2 = ['Product', 'Total Revenue (USD)', 'Avg Order (USD)', 'Total Units', 'Num Orders']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

for row_idx, row in product_stats.iterrows():
    for col_idx, value in enumerate(row, 1):
        ws2.cell(row=row_idx+4, column=col_idx, value=value)

# ── Sheet 3: Key Insights
ws3 = wb.create_sheet("Key Insights")
ws3['A1'] = "Key Business Insights"
ws3['A1'].font = Font(bold=True, size=14, color="2B86C5")

insights = [
    ("Total Revenue (USD)", f"${total_revenue:,.0f}"),
    ("Average Order Value", f"${avg_order:.0f}"),
    ("Best Performing Product", best_product),
    ("Best Performing Country", best_country),
    ("Revenue Std Deviation", f"${std_dev:,.0f}"),
    ("EUR/USD Rate Used", str(rate)),
    ("Total Orders", str(len(df))),
    ("Total Countries", str(df['country'].nunique())),
]

green_fill = PatternFill(start_color="E1F5EE", end_color="E1F5EE", fill_type="solid")
for i, (label, value) in enumerate(insights, 3):
    ws3.cell(row=i, column=1, value=label).font = Font(bold=True)
    cell = ws3.cell(row=i, column=2, value=value)
    cell.fill = green_fill
    cell.border = border

wb.save("sales_intelligence_report.xlsx")
print("Excel report saved!")

# ──────────────────────────────────────────────
# STEP 6: SAVE JSON SUMMARY
# ──────────────────────────────────────────────
print("\n[6/6] Saving JSON summary...")

summary = {
    "report_title": "Global Sales Intelligence Dashboard",
    "eur_usd_rate": rate,
    "total_orders": len(df),
    "total_revenue_usd": round(total_revenue, 2),
    "average_order_usd": round(avg_order, 2),
    "best_product": best_product,
    "best_country": best_country,
    "revenue_std_deviation": round(std_dev, 2),
    "product_breakdown": product_stats.to_dict('records'),
    "country_breakdown": country_stats.to_dict('records'),
}

with open("summary.json", "w") as f:
    json.dump(summary, f, indent=4)

print("JSON summary saved!")

print("\n" + "=" * 60)
print("CAPSTONE PROJECT COMPLETE!")
print("=" * 60)
print("Files generated:")
print("  dashboard_charts.png — 4 analysis charts")
print("  sales_intelligence_report.xlsx — 3-sheet Excel report")
print("  summary.json — key insights in JSON format")